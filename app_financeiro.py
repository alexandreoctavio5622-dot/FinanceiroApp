import streamlit as st
import openpyxl
from datetime import datetime
import re
import os

# CONFIGURAÇÃO
EXCEL_PATH = "Formulario.xlsx"
SHEET_NAME = "Janeiro-26"

st.title("Controle Financeiro Pessoal")
st.write("Adicione lançamentos de Débito ou Crédito na planilha.")

def parse_valor(valor_str):
    if not valor_str: return None
    valor_str = re.sub(r"[R$\s]", "", valor_str).replace(",", ".")
    try: return float(valor_str)
    except: return None

with st.form(key='form_lancamento'):
    data = st.date_input("Data", value=datetime.today())
    descricao = st.text_input("Descrição")
    nfp = st.text_input("NFP (opcional)")
    codigo = st.text_input("Código")
    forma_pagto = st.selectbox("Forma de Pagto.", ["débito", "crédito", "dinheiro", "VA", "cartão CEA pay"])
    debito = st.text_input("Débito Conta Corrente", value="")
    credito = st.text_input("Crédito Conta Corrente", value="")
    submit_button = st.form_submit_button(label='Adicionar Lançamento')

if submit_button:
    try:
        if not os.path.exists(EXCEL_PATH):
            st.error(f"Arquivo '{EXCEL_PATH}' não encontrado!")
        else:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            if SHEET_NAME not in wb.sheetnames:
                st.error(f"Aba '{SHEET_NAME}' não encontrada!")
            else:
                sheet = wb[SHEET_NAME]
                new_row = sheet.max_row + 1
                sheet.cell(row=new_row, column=1).value = data.strftime("%d/%m/%Y")
                sheet.cell(row=new_row, column=2).value = descricao
                sheet.cell(row=new_row, column=3).value = nfp
                sheet.cell(row=new_row, column=4).value = codigo
                sheet.cell(row=new_row, column=5).value = forma_pagto
                sheet.cell(row=new_row, column=6).value = parse_valor(debito)
                sheet.cell(row=new_row, column=7).value = parse_valor(credito)
                wb.save(EXCEL_PATH)
                st.success(f"Lançamento adicionado com sucesso!")
    except Exception as e:
        st.error(f"Erro: {e}")
